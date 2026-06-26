"""POS proforma invoice quote and export."""

from __future__ import annotations

import io
import secrets
from datetime import UTC, datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.errors import NotFoundError, ValidationError
from app.models.branch import Branch
from app.models.currency import Currency
from app.models.product import Product
from app.models.product_variant import ProductVariant
from app.schemas.pos_proforma import (
    ProformaExportRequest,
    ProformaLineIn,
    ProformaLineRead,
    ProformaQuoteResponse,
)
from app.services.accounting_service import get_accounting_settings
from app.services.cart_service import _variant_tags_map
from app.services.catalog_service import map_effective_output_tax_rates, resolve_default_variant_id
from app.services.pos_proforma_pdf_service import ProformaLocale, build_proforma_pdf_bytes
from app.services.pricing_service import get_active_sell_price
from app.utils.money import q2
from app.utils.person_name import display_person_name

_XLSX_LABELS: dict[ProformaLocale, dict[str, str]] = {
    "en": {
        "title": "Proforma invoice",
        "company": "Company",
        "branch": "Branch",
        "date": "Date / time",
        "document_no": "Document no.",
        "cashier": "Cashier",
        "logo": "LOGO",
        "product": "Product",
        "variant": "Variant",
        "qty": "Quantity",
        "unit_price": "Unit price",
        "line_total": "Line total",
        "subtotal": "Subtotal",
        "tax": "Tax",
        "total": "Total",
    },
    "ar": {
        "title": "فاتورة مبدئية",
        "company": "الشركة",
        "branch": "الفرع",
        "date": "التاريخ والوقت",
        "document_no": "رقم المستند",
        "cashier": "أمين الصندوق",
        "logo": "الشعار",
        "product": "المنتج",
        "variant": "المتغير",
        "qty": "الكمية",
        "unit_price": "سعر الوحدة",
        "line_total": "إجمالي السطر",
        "subtotal": "المجموع الفرعي",
        "tax": "الضريبة",
        "total": "الإجمالي",
    },
}


def _new_document_number() -> str:
    stamp = datetime.now(UTC).strftime("%Y%m%d%H%M%S")
    return f"PF-{stamp}-{secrets.token_hex(2).upper()}"


def _cashier_display_name(
    *,
    first_name: str | None,
    father_name: str | None,
    family_name: str | None,
    email: str | None,
) -> str:
    name = display_person_name(first_name, father_name, family_name)
    return name or (email or "").strip() or "—"


def _rtl_alignment() -> Alignment:
    return Alignment(horizontal="right", readingOrder=2)


async def _resolve_currency_code(db: AsyncSession) -> str:
    acct = await get_accounting_settings(db)
    res = await db.execute(select(Currency).where(Currency.id == acct.base_currency_id))
    cur = res.scalar_one_or_none()
    return cur.code if cur else "USD"


async def _priced_lines(
    db: AsyncSession, *, lines_in: list[ProformaLineIn]
) -> tuple[list[ProformaLineRead], Decimal, Decimal, Decimal]:
    if not lines_in:
        raise ValidationError("At least one line is required")

    product_ids = {ln.product_id for ln in lines_in}
    pres = await db.execute(select(Product).where(Product.id.in_(product_ids)))
    products = {int(p.id): p for p in pres.scalars().all()}
    missing = product_ids - set(products.keys())
    if missing:
        raise NotFoundError("Product not found", details={"product_ids": sorted(missing)})

    variant_ids = {ln.variant_id for ln in lines_in if ln.variant_id is not None}
    variants: dict[int, ProductVariant] = {}
    if variant_ids:
        vres = await db.execute(select(ProductVariant).where(ProductVariant.id.in_(variant_ids)))
        variants = {int(v.id): v for v in vres.scalars().all()}

    rates = await map_effective_output_tax_rates(db, products_by_id=products)

    resolved_vids: list[int] = []
    for ln_in in lines_in:
        vid = ln_in.variant_id
        if vid is None:
            vid = await resolve_default_variant_id(db, product_id=ln_in.product_id)
        elif vid not in variants:
            raise NotFoundError("Variant not found", details={"variant_id": vid})
        resolved_vids.append(int(vid))

    variant_tags = await _variant_tags_map(db, set(resolved_vids))

    line_reads: list[ProformaLineRead] = []
    line_bases: list[tuple[ProformaLineRead, Decimal]] = []

    for ln_in, vid in zip(lines_in, resolved_vids, strict=True):
        product = products[ln_in.product_id]

        unit_price = await get_active_sell_price(db, product_id=ln_in.product_id, variant_id=vid)
        base = q2(unit_price * Decimal(ln_in.qty))
        rate = rates.get(ln_in.product_id, Decimal("0"))
        if rate < 0:
            rate = Decimal("0")
        if rate > Decimal("1"):
            rate = Decimal("1")

        _, vlabel = variant_tags.get(vid, ([], None))

        row = ProformaLineRead(
            product_id=ln_in.product_id,
            product_name=product.name,
            product_sku=product.sku or "",
            variant_id=vid,
            variant_label=vlabel,
            qty=ln_in.qty,
            unit_price=unit_price,
            line_total=base,
            tax_rate=rate,
            line_tax_amount=Decimal("0.00"),
        )
        line_reads.append(row)
        line_bases.append((row, base))

    subtotal_net = q2(sum(b for _, b in line_bases))
    if subtotal_net <= 0:
        return line_reads, subtotal_net, Decimal("0.00"), subtotal_net

    tax_sum = Decimal("0.00")
    for row, base in line_bases:
        tax = q2(base * row.tax_rate) if base > 0 else Decimal("0.00")
        row.line_tax_amount = tax
        tax_sum += tax

    tax_total = q2(tax_sum)
    total = q2(subtotal_net + tax_total)
    return line_reads, subtotal_net, tax_total, total


async def quote_proforma(db: AsyncSession, *, lines: list[ProformaLineIn]) -> ProformaQuoteResponse:
    line_reads, subtotal, tax_total, total = await _priced_lines(db, lines_in=lines)
    currency_code = await _resolve_currency_code(db)
    return ProformaQuoteResponse(
        lines=line_reads,
        subtotal=subtotal,
        tax_total=tax_total,
        total=total,
        currency_code=currency_code,
    )


async def _branch_name(db: AsyncSession, branch_id: int | None) -> str | None:
    if branch_id is None:
        return None
    res = await db.execute(select(Branch).where(Branch.id == branch_id))
    branch = res.scalar_one_or_none()
    return branch.name if branch else None


def _line_dicts(lines: list[ProformaLineRead]) -> list[dict[str, str | int | Decimal]]:
    return [
        {
            "product_name": ln.product_name,
            "variant_label": ln.variant_label or "",
            "qty": ln.qty,
            "unit_price": ln.unit_price,
            "line_total": ln.line_total,
        }
        for ln in lines
    ]


async def export_proforma_pdf(
    db: AsyncSession,
    *,
    body: ProformaExportRequest,
    cashier_first_name: str | None = None,
    cashier_father_name: str | None = None,
    cashier_family_name: str | None = None,
    cashier_email: str | None = None,
) -> tuple[bytes, str]:
    quote = await quote_proforma(db, lines=body.lines)
    locale: ProformaLocale = body.locale if body.locale in ("ar", "en") else "ar"
    branch_name = await _branch_name(db, body.branch_id)
    document_number = _new_document_number()
    cashier_name = _cashier_display_name(
        first_name=cashier_first_name,
        father_name=cashier_father_name,
        family_name=cashier_family_name,
        email=cashier_email,
    )
    pdf_bytes = build_proforma_pdf_bytes(
        locale=locale,
        company_name=settings.COMPANY_DISPLAY_NAME,
        branch_name=branch_name,
        currency_code=quote.currency_code,
        document_number=document_number,
        cashier_name=cashier_name,
        lines=_line_dicts(quote.lines),
        subtotal=quote.subtotal,
        tax_total=quote.tax_total,
        total=quote.total,
    )
    return pdf_bytes, f"proforma-{document_number}.pdf"


async def export_proforma_xlsx(
    db: AsyncSession,
    *,
    body: ProformaExportRequest,
    cashier_first_name: str | None = None,
    cashier_father_name: str | None = None,
    cashier_family_name: str | None = None,
    cashier_email: str | None = None,
) -> tuple[bytes, str]:
    quote = await quote_proforma(db, lines=body.lines)
    locale: ProformaLocale = body.locale if body.locale in ("ar", "en") else "ar"
    labels = _XLSX_LABELS.get(locale, _XLSX_LABELS["ar"])
    branch_name = await _branch_name(db, body.branch_id)
    document_number = _new_document_number()
    cashier_name = _cashier_display_name(
        first_name=cashier_first_name,
        father_name=cashier_father_name,
        family_name=cashier_family_name,
        email=cashier_email,
    )
    rtl = _rtl_alignment()
    now = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")

    wb = Workbook()
    ws = wb.active
    ws.title = labels["title"][:31]
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True)
    meta_rows = [
        [labels["logo"], ""],
        [labels["title"], ""],
        [f"{labels['company']}: {settings.COMPANY_DISPLAY_NAME}", ""],
    ]
    if branch_name:
        meta_rows.append([f"{labels['branch']}: {branch_name}", ""])
    meta_rows.extend(
        [
            [f"{labels['date']}: {now}", ""],
            [f"{labels['document_no']}: {document_number}", ""],
            [f"{labels['cashier']}: {cashier_name}", ""],
            [],
        ]
    )
    for row in meta_rows:
        ws.append(row)
        if row:
            for cell in ws[ws.max_row]:
                cell.alignment = rtl
                if ws.max_row <= 2:
                    cell.font = header_font

    headers = [
        labels["product"],
        labels["variant"],
        labels["qty"],
        labels["unit_price"],
        labels["line_total"],
    ]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.alignment = rtl

    for ln in quote.lines:
        ws.append(
            [
                ln.product_name,
                ln.variant_label or "",
                ln.qty,
                float(ln.unit_price),
                float(ln.line_total),
            ]
        )
        for cell in ws[ws.max_row]:
            cell.alignment = rtl

    ws.append([])
    ws.append([labels["subtotal"], "", "", "", float(quote.subtotal)])
    ws.append([labels["tax"], "", "", "", float(quote.tax_total)])
    row_total = ws.max_row + 1
    ws.append([labels["total"], "", "", "", float(quote.total)])
    for row_idx in range(row_total - 2, row_total + 1):
        for cell in ws[row_idx]:
            cell.alignment = rtl
            if cell.column in (1, 5):
                cell.font = header_font

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"proforma-{document_number}.xlsx"
