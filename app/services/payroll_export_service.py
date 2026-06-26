"""Payroll period and individual payslip exports (PDF + native .xlsx)."""

from __future__ import annotations

from datetime import date

from fpdf import FPDF
from openpyxl import Workbook

from app.services.payroll_pdf_service import _fmt_paid, _register_unicode_font, _txt
from app.services.report_xlsx_utils import (
    append_meta_rows,
    cell_alignment,
    configure_sheet_locale,
    header_font,
    workbook_to_bytes,
    write_table,
)
from app.utils.request_locale import RequestLocale

PayrollLocale = RequestLocale

_PERIOD_LABELS: dict[PayrollLocale, dict[str, str]] = {
    "en": {
        "title": "Payroll sheet",
        "period": "Period",
        "employee": "Employee",
        "role": "Role",
        "base": "Base salary",
        "hourly": "Hourly rate",
        "gross": "Gross",
        "auto_ded": "Auto deductions",
        "manual_ded": "Manual deductions",
        "bonus": "Bonus",
        "overtime": "Overtime",
        "net": "Net",
        "status": "Status",
        "paid": "Paid at",
    },
    "ar": {
        "title": "مسير رواتب",
        "period": "الفترة",
        "employee": "الموظف",
        "role": "الدور",
        "base": "الراتب الأساسي",
        "hourly": "أجر الساعة",
        "gross": "الإجمالي",
        "auto_ded": "خصومات تلقائية",
        "manual_ded": "خصومات يدوية",
        "bonus": "مكافأة",
        "overtime": "عمل إضافي",
        "net": "الصافي",
        "status": "الحالة",
        "paid": "تاريخ الدفع",
    },
}

_PAYSLIP_LABELS: dict[PayrollLocale, dict[str, str]] = {
    "en": {
        "title": "Employee payslip",
        "employee": "Employee",
        "period": "Period",
        "base": "Base salary",
        "hours": "Hours worked",
        "hourly": "Hourly rate",
        "gross": "Gross amount",
        "auto_ded": "Automatic deductions",
        "manual_ded": "Manual deductions",
        "bonus": "Bonus",
        "overtime": "Overtime",
        "net": "Net pay",
        "status": "Status",
        "paid": "Paid at",
    },
    "ar": {
        "title": "قسيمة راتب",
        "employee": "الموظف",
        "period": "الفترة",
        "base": "الراتب الأساسي",
        "hours": "ساعات العمل",
        "hourly": "أجر الساعة",
        "gross": "الإجمالي",
        "auto_ded": "خصومات تلقائية",
        "manual_ded": "خصومات يدوية",
        "bonus": "مكافأة",
        "overtime": "عمل إضافي",
        "net": "صافي الراتب",
        "status": "الحالة",
        "paid": "تاريخ الدفع",
    },
}


def _labels(locale: PayrollLocale, kind: str) -> dict[str, str]:
    pool = _PERIOD_LABELS if kind == "period" else _PAYSLIP_LABELS
    return pool.get(locale, pool["ar"])


def _cell_align(locale: PayrollLocale) -> str:
    return "R" if locale == "ar" else "L"


def build_payroll_period_pdf_localized(
    *,
    period_start: date,
    period_end: date,
    rows: list[dict],
    locale: PayrollLocale = "ar",
    title: str | None = None,
) -> bytes:
    labels = _labels(locale, "period")
    rtl = locale == "ar"
    align = _cell_align(locale)
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    family = _register_unicode_font(pdf)

    pdf.set_font(family, size=12)
    pdf.cell(0, 8, _txt(title or labels["title"], 120), align=align, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(family, size=8)
    pdf.cell(
        0,
        5,
        _txt(f"{labels['period']}: {period_start.isoformat()} - {period_end.isoformat()}", 120),
        align=align,
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.ln(2)

    headers = [
        labels["employee"],
        labels["role"],
        labels["base"],
        labels["hourly"],
        labels["gross"],
        labels["auto_ded"],
        labels["manual_ded"],
        labels["bonus"],
        labels["overtime"],
        labels["net"],
        labels["status"],
        labels["paid"],
    ]
    col_w = [38, 22, 18, 18, 18, 18, 18, 16, 16, 20, 22, 28]
    if rtl:
        headers = list(reversed(headers))
        col_w = list(reversed(col_w))

    pdf.set_font(family, size=7)
    for h, w in zip(headers, col_w, strict=True):
        pdf.cell(w, 6, _txt(h, 24), border=1, align=align)
    pdf.ln()

    for r in rows:
        name = r.get("user_full_name") or r.get("user_email") or str(r["employee_profile_id"])
        line = [
            name,
            r.get("user_role_code"),
            r.get("base_salary"),
            r.get("hourly_rate"),
            r.get("gross_amount"),
            r.get("automatic_deductions_amount"),
            r.get("manual_deductions_amount"),
            r.get("bonus_amount"),
            r.get("overtime_amount"),
            r.get("net_amount"),
            r.get("payslip_status"),
            _fmt_paid(r.get("paid_at")),
        ]
        if rtl:
            line = list(reversed(line))
        for val, w in zip(line, col_w, strict=True):
            pdf.cell(w, 6, _txt(val, 48), border=1, align=align)
        pdf.ln()

    return bytes(pdf.output())


def build_payroll_period_xlsx(
    *,
    period_start: date,
    period_end: date,
    rows: list[dict],
    locale: PayrollLocale = "ar",
) -> bytes:
    labels = _labels(locale, "period")
    wb = Workbook()
    ws = wb.active
    ws.title = labels["title"][:31]
    configure_sheet_locale(ws, locale)

    append_meta_rows(
        ws,
        [
            (labels["title"], ""),
            (labels["period"], f"{period_start.isoformat()} - {period_end.isoformat()}"),
        ],
        locale=locale,
        title_row=True,
    )
    ws.append([])

    table_rows: list[list[object]] = []
    for r in rows:
        name = (
            r.get("user_full_name") or r.get("user_email") or str(r.get("employee_profile_id", ""))
        )
        table_rows.append(
            [
                name,
                r.get("user_role_code"),
                r.get("base_salary"),
                r.get("hourly_rate"),
                r.get("gross_amount"),
                r.get("automatic_deductions_amount"),
                r.get("manual_deductions_amount"),
                r.get("bonus_amount"),
                r.get("overtime_amount"),
                r.get("net_amount"),
                r.get("payslip_status"),
                r.get("paid_at"),
            ]
        )
    write_table(
        ws,
        [
            labels["employee"],
            labels["role"],
            labels["base"],
            labels["hourly"],
            labels["gross"],
            labels["auto_ded"],
            labels["manual_ded"],
            labels["bonus"],
            labels["overtime"],
            labels["net"],
            labels["status"],
            labels["paid"],
        ],
        table_rows,
        locale=locale,
    )
    return workbook_to_bytes(wb)


def build_payslip_pdf(
    payslip: dict,
    *,
    locale: PayrollLocale = "ar",
) -> bytes:
    labels = _labels(locale, "payslip")
    align = _cell_align(locale)
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    family = _register_unicode_font(pdf)
    page_w = pdf.w - pdf.l_margin - pdf.r_margin

    name = (
        payslip.get("user_full_name")
        or payslip.get("user_email")
        or str(payslip.get("employee_profile_id", ""))
    )
    period = f"{payslip.get('period_start')} - {payslip.get('period_end')}"

    pdf.set_font(family, size=14)
    pdf.cell(page_w, 8, _txt(labels["title"], 80), align=align, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(family, size=10)
    for label, value in (
        (labels["employee"], name),
        (labels["period"], period),
        (labels["base"], payslip.get("base_salary_amount")),
        (labels["hours"], payslip.get("hours_worked")),
        (labels["hourly"], payslip.get("hourly_rate")),
        (labels["gross"], payslip.get("gross_amount")),
        (labels["auto_ded"], payslip.get("automatic_deductions_amount")),
        (labels["manual_ded"], payslip.get("manual_deductions_amount")),
        (labels["bonus"], payslip.get("bonus_amount")),
        (labels["overtime"], payslip.get("overtime_amount")),
        (labels["net"], payslip.get("net_amount")),
        (labels["status"], payslip.get("status")),
        (labels["paid"], _fmt_paid(payslip.get("paid_at"))),
    ):
        pdf.cell(
            page_w, 6, _txt(f"{label}: {value}", 100), align=align, new_x="LMARGIN", new_y="NEXT"
        )

    return bytes(pdf.output())


def build_payslip_xlsx(
    payslip: dict,
    *,
    locale: PayrollLocale = "ar",
) -> bytes:
    labels = _labels(locale, "payslip")
    wb = Workbook()
    ws = wb.active
    ws.title = labels["title"][:31]
    configure_sheet_locale(ws, locale)

    name = (
        payslip.get("user_full_name")
        or payslip.get("user_email")
        or str(payslip.get("employee_profile_id", ""))
    )
    period = f"{payslip.get('period_start')} - {payslip.get('period_end')}"

    rows = [
        (labels["title"], ""),
        (labels["employee"], name),
        (labels["period"], period),
        (labels["base"], payslip.get("base_salary_amount")),
        (labels["hours"], payslip.get("hours_worked")),
        (labels["hourly"], payslip.get("hourly_rate")),
        (labels["gross"], payslip.get("gross_amount")),
        (labels["auto_ded"], payslip.get("automatic_deductions_amount")),
        (labels["manual_ded"], payslip.get("manual_deductions_amount")),
        (labels["bonus"], payslip.get("bonus_amount")),
        (labels["overtime"], payslip.get("overtime_amount")),
        (labels["net"], payslip.get("net_amount")),
        (labels["status"], payslip.get("status")),
        (labels["paid"], payslip.get("paid_at")),
    ]
    append_meta_rows(ws, rows, locale=locale, title_row=True)
    for row_idx in range(1, ws.max_row + 1):
        if row_idx <= 2:
            for cell in ws[row_idx]:
                cell.font = header_font()
        for cell in ws[row_idx]:
            cell.alignment = cell_alignment(locale)
    return workbook_to_bytes(wb)


def payslip_dict_from_read(payslip_read) -> dict:
    """Convert PayslipRead (or compatible) to a plain dict for export builders."""
    if hasattr(payslip_read, "model_dump"):
        return payslip_read.model_dump()
    return dict(payslip_read)
