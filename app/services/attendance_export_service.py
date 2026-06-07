"""Attendance and leave summary exports (PDF + native .xlsx)."""

from __future__ import annotations

from datetime import date

from fpdf import FPDF
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.branch import Branch
from app.models.employee_profile import EmployeeProfile
from app.models.leave_request import LeaveRequest
from app.models.users import User
from app.services.employee_service import (
    attendance_period_summary,
    list_attendance_logs_enriched_page,
    list_leave_requests_filtered,
)
from app.services.payroll_pdf_service import _register_unicode_font, _txt
from app.services.report_xlsx_utils import (
    append_meta_rows,
    configure_sheet_locale,
    workbook_to_bytes,
    write_table,
)
from app.utils.person_name import person_name_sql_expr
from app.utils.request_locale import RequestLocale

HrLocale = RequestLocale

_LABELS: dict[HrLocale, dict[str, str]] = {
    "en": {
        "attendance_title": "Attendance summary",
        "leave_title": "Leave requests summary",
        "period": "Period",
        "branch": "Branch",
        "employee": "Employee",
        "clock_in": "Clock in",
        "clock_out": "Clock out",
        "status": "Status",
        "category": "Category",
        "overtime_min": "Overtime (min)",
        "record_count": "Records",
        "absent_days": "Absent days",
        "leave_type": "Leave type",
        "start_date": "Start",
        "end_date": "End",
        "days": "Days",
        "review_status": "Status",
    },
    "ar": {
        "attendance_title": "ملخص الحضور والانصراف",
        "leave_title": "ملخص الإجازات",
        "period": "الفترة",
        "branch": "الفرع",
        "employee": "الموظف",
        "clock_in": "تسجيل الدخول",
        "clock_out": "تسجيل الخروج",
        "status": "الحالة",
        "category": "الفئة",
        "overtime_min": "عمل إضافي (د)",
        "record_count": "السجلات",
        "absent_days": "أيام الغياب",
        "leave_type": "نوع الإجازة",
        "start_date": "البداية",
        "end_date": "النهاية",
        "days": "الأيام",
        "review_status": "الحالة",
    },
}


def _labels(locale: HrLocale) -> dict[str, str]:
    return _LABELS.get(locale, _LABELS["ar"])


def _cell_align(locale: HrLocale) -> str:
    return "R" if locale == "ar" else "L"


async def _branch_name(db: AsyncSession, branch_id: int | None) -> str | None:
    if branch_id is None:
        return None
    res = await db.execute(select(Branch.name).where(Branch.id == branch_id))
    val = res.scalar_one_or_none()
    return str(val) if val else None


async def export_attendance_pdf(
    db: AsyncSession,
    *,
    branch_id: int | None = None,
    employee_profile_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    locale: HrLocale = "ar",
) -> tuple[bytes, str]:
    labels = _labels(locale)
    align = _cell_align(locale)
    summary = await attendance_period_summary(
        db,
        branch_id=branch_id,
        employee_profile_id=employee_profile_id,
        date_from=date_from,
        date_to=date_to,
    )
    logs, _total = await list_attendance_logs_enriched_page(
        db,
        branch_id=branch_id,
        employee_profile_id=employee_profile_id,
        date_from=date_from,
        date_to=date_to,
        limit=500,
        offset=0,
    )

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    family = _register_unicode_font(pdf)
    pdf.set_font(family, size=12)
    pdf.cell(0, 8, _txt(labels["attendance_title"], 120), align=align, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(family, size=8)
    branch_name = await _branch_name(db, branch_id)
    if branch_name:
        pdf.cell(0, 5, _txt(f"{labels['branch']}: {branch_name}", 80), align=align, new_x="LMARGIN", new_y="NEXT")
    if date_from and date_to:
        pdf.cell(
            0,
            5,
            _txt(f"{labels['period']}: {date_from.isoformat()} - {date_to.isoformat()}", 80),
            align=align,
            new_x="LMARGIN",
            new_y="NEXT",
        )
    pdf.cell(
        0,
        5,
        _txt(
            f"{labels['record_count']}: {summary['record_count']} | "
            f"{labels['absent_days']}: {summary['absent_days']}",
            120,
        ),
        align=align,
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.ln(2)

    headers = [
        labels["employee"],
        labels["clock_in"],
        labels["clock_out"],
        labels["category"],
        labels["status"],
        labels["overtime_min"],
    ]
    col_w = [44, 36, 36, 28, 28, 28]
    if locale == "ar":
        headers = list(reversed(headers))
        col_w = list(reversed(col_w))

    pdf.set_font(family, size=7)
    for h, w in zip(headers, col_w, strict=True):
        pdf.cell(w, 6, _txt(h, 24), border=1, align=align)
    pdf.ln()

    for log in logs:
        name = log.get("employee_user_full_name") or log.get("employee_user_email") or str(
            log.get("employee_profile_id", "")
        )
        row = [
            name,
            log.get("clock_in_at"),
            log.get("clock_out_at") or "—",
            log.get("attendance_category"),
            log.get("classification_status"),
            log.get("overtime_minutes"),
        ]
        if locale == "ar":
            row = list(reversed(row))
        for val, w in zip(row, col_w, strict=True):
            pdf.cell(w, 6, _txt(val, 48), border=1, align=align)
        pdf.ln()

    period = f"{date_from or 'all'}-{date_to or 'all'}"
    return bytes(pdf.output()), f"attendance-{period}.pdf"


async def export_attendance_xlsx(
    db: AsyncSession,
    *,
    branch_id: int | None = None,
    employee_profile_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    locale: HrLocale = "ar",
) -> tuple[bytes, str]:
    labels = _labels(locale)
    summary = await attendance_period_summary(
        db,
        branch_id=branch_id,
        employee_profile_id=employee_profile_id,
        date_from=date_from,
        date_to=date_to,
    )
    logs, _total = await list_attendance_logs_enriched_page(
        db,
        branch_id=branch_id,
        employee_profile_id=employee_profile_id,
        date_from=date_from,
        date_to=date_to,
        limit=500,
        offset=0,
    )
    branch_name = await _branch_name(db, branch_id)

    wb = Workbook()
    ws = wb.active
    ws.title = labels["attendance_title"][:31]
    configure_sheet_locale(ws, locale)
    meta = [(labels["attendance_title"], "")]
    if branch_name:
        meta.append((labels["branch"], branch_name))
    if date_from and date_to:
        meta.append((labels["period"], f"{date_from.isoformat()} - {date_to.isoformat()}"))
    meta.append((labels["record_count"], str(summary["record_count"])))
    meta.append((labels["absent_days"], str(summary["absent_days"])))
    append_meta_rows(ws, meta, locale=locale, title_row=True)
    ws.append([])

    table_rows = []
    for log in logs:
        name = log.get("employee_user_full_name") or log.get("employee_user_email") or str(
            log.get("employee_profile_id", "")
        )
        table_rows.append(
            [
                name,
                log.get("clock_in_at"),
                log.get("clock_out_at"),
                log.get("attendance_category"),
                log.get("classification_status"),
                log.get("overtime_minutes"),
            ]
        )
    write_table(
        ws,
        [
            labels["employee"],
            labels["clock_in"],
            labels["clock_out"],
            labels["category"],
            labels["status"],
            labels["overtime_min"],
        ],
        table_rows,
        locale=locale,
    )
    period = f"{date_from or 'all'}-{date_to or 'all'}"
    return workbook_to_bytes(wb), f"attendance-{period}.xlsx"


async def export_leave_summary_pdf(
    db: AsyncSession,
    *,
    status: str | None = None,
    employee_profile_id: int | None = None,
    locale: HrLocale = "ar",
) -> tuple[bytes, str]:
    labels = _labels(locale)
    align = _cell_align(locale)
    leaves = await list_leave_requests_filtered(
        db,
        status=status,
        employee_profile_id=employee_profile_id,
        limit=500,
        offset=0,
    )
    employee_name = person_name_sql_expr(User.first_name, User.father_name, User.family_name)
    rows_data: list[list[object]] = []
    for leave in leaves:
        res = await db.execute(
            select(employee_name, User.email)
            .join(EmployeeProfile, EmployeeProfile.user_id == User.id)
            .where(EmployeeProfile.id == leave.employee_profile_id)
        )
        emp = res.one_or_none()
        name = emp[0] if emp and emp[0] else (emp[1] if emp else str(leave.employee_profile_id))
        day_span = (leave.end_date - leave.start_date).days + 1
        rows_data.append(
            [
                name,
                leave.leave_type.value if hasattr(leave.leave_type, "value") else str(leave.leave_type),
                leave.start_date.isoformat(),
                leave.end_date.isoformat(),
                day_span,
                leave.status.value if hasattr(leave.status, "value") else str(leave.status),
            ]
        )

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    family = _register_unicode_font(pdf)
    pdf.set_font(family, size=12)
    pdf.cell(0, 8, _txt(labels["leave_title"], 120), align=align, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    headers = [
        labels["employee"],
        labels["leave_type"],
        labels["start_date"],
        labels["end_date"],
        labels["days"],
        labels["review_status"],
    ]
    col_w = [44, 28, 28, 28, 18, 28]
    if locale == "ar":
        headers = list(reversed(headers))
        col_w = list(reversed(col_w))

    pdf.set_font(family, size=7)
    for h, w in zip(headers, col_w, strict=True):
        pdf.cell(w, 6, _txt(h, 24), border=1, align=align)
    pdf.ln()
    for row in rows_data:
        if locale == "ar":
            row = list(reversed(row))
        for val, w in zip(row, col_w, strict=True):
            pdf.cell(w, 6, _txt(val, 48), border=1, align=align)
        pdf.ln()

    return bytes(pdf.output()), "leave-summary.pdf"


async def export_leave_summary_xlsx(
    db: AsyncSession,
    *,
    status: str | None = None,
    employee_profile_id: int | None = None,
    locale: HrLocale = "ar",
) -> tuple[bytes, str]:
    labels = _labels(locale)
    leaves = await list_leave_requests_filtered(
        db,
        status=status,
        employee_profile_id=employee_profile_id,
        limit=500,
        offset=0,
    )
    employee_name = person_name_sql_expr(User.first_name, User.father_name, User.family_name)
    table_rows: list[list[object]] = []
    for leave in leaves:
        res = await db.execute(
            select(employee_name, User.email)
            .join(EmployeeProfile, EmployeeProfile.user_id == User.id)
            .where(EmployeeProfile.id == leave.employee_profile_id)
        )
        emp = res.one_or_none()
        name = emp[0] if emp and emp[0] else (emp[1] if emp else str(leave.employee_profile_id))
        day_span = (leave.end_date - leave.start_date).days + 1
        table_rows.append(
            [
                name,
                leave.leave_type.value if hasattr(leave.leave_type, "value") else str(leave.leave_type),
                leave.start_date.isoformat(),
                leave.end_date.isoformat(),
                day_span,
                leave.status.value if hasattr(leave.status, "value") else str(leave.status),
            ]
        )

    wb = Workbook()
    ws = wb.active
    ws.title = labels["leave_title"][:31]
    configure_sheet_locale(ws, locale)
    append_meta_rows(ws, [(labels["leave_title"], "")], locale=locale, title_row=True)
    ws.append([])
    write_table(
        ws,
        [
            labels["employee"],
            labels["leave_type"],
            labels["start_date"],
            labels["end_date"],
            labels["days"],
            labels["review_status"],
        ],
        table_rows,
        locale=locale,
    )
    return workbook_to_bytes(wb), "leave-summary.xlsx"
