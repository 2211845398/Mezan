"""Shared helpers for native .xlsx report exports with RTL/LTR layout."""

from __future__ import annotations

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.request_locale import RequestLocale


def cell_alignment(locale: RequestLocale) -> Alignment:
    if locale == "ar":
        return Alignment(horizontal="right", readingOrder=2)
    return Alignment(horizontal="left", readingOrder=1)


def header_font() -> Font:
    return Font(bold=True)


def configure_sheet_locale(ws: Worksheet, locale: RequestLocale) -> None:
    ws.sheet_view.rightToLeft = locale == "ar"


def append_meta_rows(
    ws: Worksheet,
    rows: Iterable[tuple[str, str]],
    *,
    locale: RequestLocale,
    title_row: bool = False,
) -> None:
    for idx, (label, value) in enumerate(rows):
        ws.append([f"{label}: {value}"])
        for cell in ws[ws.max_row]:
            cell.alignment = cell_alignment(locale)
            if title_row and idx == 0:
                cell.font = header_font()


def write_table(
    ws: Worksheet,
    headers: list[str],
    rows: list[list[object]],
    *,
    locale: RequestLocale,
) -> None:
    hdrs = list(reversed(headers)) if locale == "ar" else headers
    ws.append(hdrs)
    for cell in ws[ws.max_row]:
        cell.font = header_font()
        cell.alignment = cell_alignment(locale)
    for row in rows:
        data = list(reversed(row)) if locale == "ar" else row
        ws.append(data)
        for cell in ws[ws.max_row]:
            cell.alignment = cell_alignment(locale)


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
