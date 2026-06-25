"""Sales invoice, register, daily summary, and credit note exports (PDF + .xlsx)."""

from __future__ import annotations

from collections import defaultdict
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.errors import NotFoundError
from app.models.branch import Branch
from app.models.customer_profile import CustomerProfile
from app.models.product import Product
from app.models.product_variant import ProductVariant
from app.models.sales_invoice import SalesInvoice, SalesInvoiceLine
from app.models.sales_return import CreditNote, SalesReturn, SalesReturnLine
from app.schemas.sales_invoice import SalesInvoiceListItem
from app.services.invoice_service import list_sales_invoices_register_page, read_sales_invoice_detail
from app.services.payroll_pdf_service import _register_unicode_font, _txt
from app.services.report_xlsx_utils import (
    append_meta_rows,
    configure_sheet_locale,
    header_font,
    cell_alignment,
    workbook_to_bytes,
    write_table,
)
from app.utils.person_name import display_person_name
from app.utils.request_locale import RequestLocale

SalesLocale = RequestLocale

_LABELS: dict[SalesLocale, dict[str, str]] = {
    "en": {
        "invoice_title": "Sales invoice",
        "credit_title": "Credit note",
        "register_title": "Sales invoice register",
        "daily_title": "Daily sales summary",
        "company": "Company",
        "branch": "Branch",
        "date": "Date",
        "time": "Time",
        "document_no": "Document no.",
        "customer": "Customer",
        "logo": "LOGO",
        "product": "Product",
        "variant": "Variant",
        "qty": "Qty",
        "unit_price": "Unit price",
        "line_total": "Line total",
        "subtotal": "Subtotal",
        "discount": "Discount",
        "tax": "Tax",
        "total": "Total",
        "invoice_number": "Invoice no.",
        "txn_type": "Type",
        "payment_status": "Payment",
        "sale": "Sale",
        "return": "Return",
        "day": "Day",
        "count": "Count",
        "sales_total": "Sales total",
        "returns_total": "Returns total",
        "net_total": "Net total",
        "none": "—",
        "period": "Period",
    },
    "ar": {
        "invoice_title": "فاتورة مبيعات",
        "credit_title": "إشعار دائن",
        "register_title": "سجل فواتير المبيعات",
        "daily_title": "ملخص المبيعات اليومي",
        "company": "الشركة",
        "branch": "الفرع",
        "date": "التاريخ",
        "time": "الوقت",
        "document_no": "رقم المستند",
        "customer": "العميل",
        "logo": "الشعار",
        "product": "المنتج",
        "variant": "المتغير",
        "qty": "الكمية",
        "unit_price": "سعر الوحدة",
        "line_total": "إجمالي السطر",
        "subtotal": "المجموع الفرعي",
        "discount": "الخصم",
        "tax": "الضريبة",
        "total": "الإجمالي",
        "invoice_number": "رقم الفاتورة",
        "txn_type": "النوع",
        "payment_status": "الدفع",
        "sale": "بيع",
        "return": "مرتجع",
        "day": "اليوم",
        "count": "العدد",
        "sales_total": "إجمالي المبيعات",
        "returns_total": "إجمالي المرتجعات",
        "net_total": "الصافي",
        "none": "—",
        "period": "الفترة",
    },
}


def _labels(locale: SalesLocale) -> dict[str, str]:
    return _LABELS.get(locale, _LABELS["ar"])


def _fmt_money(val: Decimal | str | None) -> str:
    if val is None:
        return "0.00"
    return f"{Decimal(str(val)):.2f}"


def _cell_align(locale: SalesLocale) -> str:
    return "R" if locale == "ar" else "L"


def _resolve_mezan_logo_path() -> Path | None:
    root = Path(__file__).resolve().parents[2]
    for rel in (
        "app/assets/branding/mezan_logo.png",
        "mobile/web/icons/Icon-192.png",
        "mobile/web/favicon.png",
    ):
        candidate = root / rel
        if candidate.is_file():
            return candidate
    return None


def _draw_pdf_logo_banner(pdf: FPDF, *, locale: SalesLocale, family: str, labels: dict[str, str]) -> None:
    rtl = locale == "ar"
    logo_w, logo_h = 28.0, 18.0
    logo_x = pdf.l_margin if rtl else pdf.w - pdf.r_margin - logo_w
    y0 = pdf.get_y()
    logo_path = _resolve_mezan_logo_path()
    if logo_path is not None:
        try:
            pdf.image(str(logo_path), x=logo_x, y=y0, w=logo_w, h=logo_h)
            pdf.set_y(y0 + logo_h + 2)
            return
        except Exception:
            pass
    pdf.set_draw_color(180, 180, 180)
    pdf.rect(logo_x, y0, logo_w, logo_h)
    pdf.set_font(family, size=8)
    pdf.set_xy(logo_x, y0 + logo_h / 2 - 2)
    pdf.cell(logo_w, 4, _txt(labels["logo"], 12), align="C")
    pdf.set_y(y0 + logo_h + 2)


def _register_datetime_parts(dt: datetime) -> tuple[str, str]:
    aware = dt if dt.tzinfo is not None else dt.replace(tzinfo=UTC)
    utc = aware.astimezone(UTC)
    return utc.strftime("%Y-%m-%d"), utc.strftime("%H:%M")


def _build_document_pdf(
    *,
    locale: SalesLocale,
    title: str,
    company_name: str,
    branch_name: str | None,
    currency_code: str,
    document_number: str,
    customer_name: str | None,
    created_at: datetime,
    lines: list[dict[str, str | int | Decimal]],
    subtotal: Decimal,
    discount_total: Decimal,
    tax_total: Decimal,
    total: Decimal,
) -> bytes:
    labels = _labels(locale)
    rtl = locale == "ar"
    align = _cell_align(locale)

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    family = _register_unicode_font(pdf)
    page_w = pdf.w - pdf.l_margin - pdf.r_margin

    logo_w, logo_h = 28.0, 18.0
    logo_x = pdf.l_margin if rtl else pdf.w - pdf.r_margin - logo_w
    y0 = pdf.get_y()
    logo_path = _resolve_mezan_logo_path()
    if logo_path is not None:
        try:
            pdf.image(str(logo_path), x=logo_x, y=y0, w=logo_w, h=logo_h)
            pdf.set_y(y0 + logo_h + 2)
        except Exception:
            pdf.set_draw_color(180, 180, 180)
            pdf.rect(logo_x, y0, logo_w, logo_h)
            pdf.set_font(family, size=8)
            pdf.set_xy(logo_x, y0 + logo_h / 2 - 2)
            pdf.cell(logo_w, 4, _txt(labels["logo"], 12), align="C")
            pdf.set_y(y0 + logo_h + 2)
    else:
        pdf.set_draw_color(180, 180, 180)
        pdf.rect(logo_x, y0, logo_w, logo_h)
        pdf.set_font(family, size=8)
        pdf.set_xy(logo_x, y0 + logo_h / 2 - 2)
        pdf.cell(logo_w, 4, _txt(labels["logo"], 12), align="C")
        pdf.set_y(y0 + logo_h + 2)

    pdf.set_font(family, size=14)
    pdf.cell(page_w, 8, _txt(title, 80), align=align, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(family, size=9)
    meta = [
        (labels["company"], company_name),
        (labels["branch"], branch_name or labels["none"]),
        (labels["date"], created_at.strftime("%Y-%m-%d %H:%M UTC")),
        (labels["document_no"], document_number),
    ]
    if customer_name:
        meta.append((labels["customer"], customer_name))
    for label, value in meta:
        pdf.cell(page_w, 5, _txt(f"{label}: {value}", 120), align=align, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    col_w = [52, 38, 18, 28, 28]
    headers = [labels["product"], labels["variant"], labels["qty"], labels["unit_price"], labels["line_total"]]
    if rtl:
        col_w = list(reversed(col_w))
        headers = list(reversed(headers))

    pdf.set_font(family, size=8)
    for w, h in zip(col_w, headers, strict=True):
        pdf.cell(w, 6, _txt(h, 28), border=1, align=align)
    pdf.ln()

    for row in lines:
        variant = str(row.get("variant_label") or "").strip() or labels["none"]
        cells = [
            str(row["product_name"]),
            variant,
            str(row["qty"]),
            _fmt_money(row["unit_price"]),
            _fmt_money(row["line_total"]),
        ]
        if rtl:
            cells = list(reversed(cells))
        for val, w in zip(cells, col_w, strict=True):
            pdf.cell(w, 5, _txt(val, 30), border=1, align=align)
        pdf.ln()

    pdf.ln(2)
    summary_w = page_w * 0.55
    pdf.set_x(pdf.l_margin if rtl else pdf.w - pdf.r_margin - summary_w)
    pdf.set_font(family, size=9)
    for label, amount in (
        (labels["subtotal"], subtotal),
        (labels["discount"], discount_total),
        (labels["tax"], tax_total),
    ):
        pdf.cell(
            summary_w,
            5,
            _txt(f"{label}: {_fmt_money(amount)} {currency_code}", 60),
            align=align,
            new_x="LMARGIN",
            new_y="NEXT",
        )
        if not rtl:
            pdf.set_x(pdf.w - pdf.r_margin - summary_w)
    pdf.set_font(family, size=10)
    pdf.cell(
        summary_w,
        6,
        _txt(f"{labels['total']}: {_fmt_money(total)} {currency_code}", 60),
        align=align,
        new_x="LMARGIN",
        new_y="NEXT",
    )
    return bytes(pdf.output())


def _build_document_xlsx(
    *,
    locale: SalesLocale,
    title: str,
    company_name: str,
    branch_name: str | None,
    currency_code: str,
    document_number: str,
    customer_name: str | None,
    created_at: datetime,
    lines: list[dict[str, str | int | Decimal]],
    subtotal: Decimal,
    discount_total: Decimal,
    tax_total: Decimal,
    total: Decimal,
) -> bytes:
    labels = _labels(locale)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    configure_sheet_locale(ws, locale)

    meta = [
        (labels["company"], company_name),
        (title, ""),
        (labels["branch"], branch_name or labels["none"]),
        (labels["date"], created_at.strftime("%Y-%m-%d %H:%M UTC")),
        (labels["document_no"], document_number),
    ]
    if customer_name:
        meta.append((labels["customer"], customer_name))
    append_meta_rows(ws, meta, locale=locale, title_row=True)
    ws.append([])

    table_rows = [
        [
            ln["product_name"],
            ln.get("variant_label") or "",
            ln["qty"],
            float(Decimal(str(ln["unit_price"]))),
            float(Decimal(str(ln["line_total"]))),
        ]
        for ln in lines
    ]
    write_table(
        ws,
        [labels["product"], labels["variant"], labels["qty"], labels["unit_price"], labels["line_total"]],
        table_rows,
        locale=locale,
    )
    ws.append([])
    for label, amount in (
        (labels["subtotal"], subtotal),
        (labels["discount"], discount_total),
        (labels["tax"], tax_total),
        (labels["total"], total),
    ):
        ws.append([label, "", "", "", float(Decimal(str(amount)))])
        for cell in ws[ws.max_row]:
            cell.alignment = cell_alignment(locale)
            if cell.column in (1, 5):
                cell.font = header_font()
    _ = currency_code
    return workbook_to_bytes(wb)


async def export_sales_invoice_pdf(
    db: AsyncSession,
    *,
    invoice_id: int,
    locale: SalesLocale = "ar",
) -> tuple[bytes, str]:
    detail = await read_sales_invoice_detail(db, invoice_id=invoice_id)
    labels = _labels(locale)
    company = detail.company_legal_name or settings.COMPANY_DISPLAY_NAME
    lines = [
        {
            "product_name": ln.product_name,
            "variant_label": ln.product_sku,
            "qty": ln.qty,
            "unit_price": ln.unit_price,
            "line_total": ln.line_total,
        }
        for ln in detail.lines
    ]
    pdf = _build_document_pdf(
        locale=locale,
        title=labels["invoice_title"],
        company_name=company,
        branch_name=detail.branch_name,
        currency_code=detail.currency_code or "USD",
        document_number=detail.invoice_number,
        customer_name=detail.customer_display,
        created_at=detail.created_at,
        lines=lines,
        subtotal=detail.subtotal,
        discount_total=detail.discount_total,
        tax_total=detail.tax_total,
        total=detail.total,
    )
    return pdf, f"invoice-{detail.invoice_number}.pdf"


async def export_sales_invoice_xlsx(
    db: AsyncSession,
    *,
    invoice_id: int,
    locale: SalesLocale = "ar",
) -> tuple[bytes, str]:
    detail = await read_sales_invoice_detail(db, invoice_id=invoice_id)
    labels = _labels(locale)
    company = detail.company_legal_name or settings.COMPANY_DISPLAY_NAME
    lines = [
        {
            "product_name": ln.product_name,
            "variant_label": ln.product_sku,
            "qty": ln.qty,
            "unit_price": ln.unit_price,
            "line_total": ln.line_total,
        }
        for ln in detail.lines
    ]
    xlsx = _build_document_xlsx(
        locale=locale,
        title=labels["invoice_title"],
        company_name=company,
        branch_name=detail.branch_name,
        currency_code=detail.currency_code or "USD",
        document_number=detail.invoice_number,
        customer_name=detail.customer_display,
        created_at=detail.created_at,
        lines=lines,
        subtotal=detail.subtotal,
        discount_total=detail.discount_total,
        tax_total=detail.tax_total,
        total=detail.total,
    )
    return xlsx, f"invoice-{detail.invoice_number}.xlsx"


async def _fetch_credit_note_bundle(db: AsyncSession, credit_note_id: int):
    res = await db.execute(
        select(CreditNote, SalesReturn, SalesInvoice, Branch, CustomerProfile)
        .join(SalesReturn, CreditNote.sales_return_id == SalesReturn.id)
        .join(SalesInvoice, SalesReturn.sales_invoice_id == SalesInvoice.id)
        .join(Branch, SalesInvoice.branch_id == Branch.id)
        .outerjoin(CustomerProfile, SalesInvoice.customer_id == CustomerProfile.id)
        .where(CreditNote.id == credit_note_id)
    )
    row = res.one_or_none()
    if row is None:
        raise NotFoundError("Credit note not found", details={"credit_note_id": credit_note_id})
    return row


async def export_credit_note_pdf(
    db: AsyncSession,
    *,
    credit_note_id: int,
    locale: SalesLocale = "ar",
) -> tuple[bytes, str]:
    cn, _ret, _inv, branch, cust = await _fetch_credit_note_bundle(db, credit_note_id)
    labels = _labels(locale)
    cust_disp = None
    if cust is not None:
        cust_disp = display_person_name(cust.first_name, cust.father_name, cust.family_name) or cust.phone

    line_res = await db.execute(
        select(SalesReturnLine, SalesInvoiceLine, Product, ProductVariant)
        .join(SalesInvoiceLine, SalesReturnLine.sales_invoice_line_id == SalesInvoiceLine.id)
        .join(Product, SalesReturnLine.product_id == Product.id)
        .join(ProductVariant, SalesReturnLine.variant_id == ProductVariant.id)
        .where(SalesReturnLine.sales_return_id == cn.sales_return_id)
    )
    lines = []
    for ret_ln, inv_ln, prod, variant in line_res.all():
        lines.append(
            {
                "product_name": prod.name,
                "variant_label": variant.sku,
                "qty": ret_ln.qty,
                "unit_price": inv_ln.unit_price,
                "line_total": ret_ln.refund_amount,
            }
        )

    pdf = _build_document_pdf(
        locale=locale,
        title=labels["credit_title"],
        company_name=settings.COMPANY_DISPLAY_NAME,
        branch_name=branch.name,
        currency_code="USD",
        document_number=cn.credit_number,
        customer_name=cust_disp,
        created_at=cn.created_at,
        lines=lines,
        subtotal=cn.total_amount,
        discount_total=Decimal("0"),
        tax_total=Decimal("0"),
        total=cn.total_amount,
    )
    return pdf, f"credit-note-{cn.credit_number}.pdf"


async def export_credit_note_xlsx(
    db: AsyncSession,
    *,
    credit_note_id: int,
    locale: SalesLocale = "ar",
) -> tuple[bytes, str]:
    cn, _ret, _inv, branch, cust = await _fetch_credit_note_bundle(db, credit_note_id)
    labels = _labels(locale)
    cust_disp = None
    if cust is not None:
        cust_disp = display_person_name(cust.first_name, cust.father_name, cust.family_name) or cust.phone

    line_res = await db.execute(
        select(SalesReturnLine, SalesInvoiceLine, Product, ProductVariant)
        .join(SalesInvoiceLine, SalesReturnLine.sales_invoice_line_id == SalesInvoiceLine.id)
        .join(Product, SalesReturnLine.product_id == Product.id)
        .join(ProductVariant, SalesReturnLine.variant_id == ProductVariant.id)
        .where(SalesReturnLine.sales_return_id == cn.sales_return_id)
    )
    lines = []
    for ret_ln, inv_ln, prod, variant in line_res.all():
        lines.append(
            {
                "product_name": prod.name,
                "variant_label": variant.sku,
                "qty": ret_ln.qty,
                "unit_price": inv_ln.unit_price,
                "line_total": ret_ln.refund_amount,
            }
        )

    xlsx = _build_document_xlsx(
        locale=locale,
        title=labels["credit_title"],
        company_name=settings.COMPANY_DISPLAY_NAME,
        branch_name=branch.name,
        currency_code="USD",
        document_number=cn.credit_number,
        customer_name=cust_disp,
        created_at=cn.created_at,
        lines=lines,
        subtotal=cn.total_amount,
        discount_total=Decimal("0"),
        tax_total=Decimal("0"),
        total=cn.total_amount,
    )
    return xlsx, f"credit-note-{cn.credit_number}.xlsx"


async def _register_rows_all(
    db: AsyncSession,
    *,
    branch_id: int,
    start_inclusive: datetime,
    end_exclusive: datetime,
    limit: int | None = None,
    offset: int = 0,
) -> tuple[list[SalesInvoiceListItem], str | None]:
    page_limit = limit if limit is not None else 10_000
    items, _total, _sub, _tot = await list_sales_invoices_register_page(
        db,
        branch_id=branch_id,
        start_inclusive=start_inclusive,
        end_exclusive=end_exclusive,
        limit=page_limit,
        offset=offset,
    )
    bres = await db.execute(select(Branch.name).where(Branch.id == branch_id))
    branch_name = bres.scalar_one_or_none()
    return items, str(branch_name) if branch_name else None


def _register_table_rows(items: list[SalesInvoiceListItem], locale: SalesLocale) -> list[list[object]]:
    labels = _labels(locale)
    rows: list[list[object]] = []
    for item in items:
        txn = labels["return"] if item.transaction_type == "return" else labels["sale"]
        date_str, time_str = _register_datetime_parts(item.created_at)
        rows.append(
            [
                item.invoice_number,
                item.customer_display or labels["none"],
                date_str,
                time_str,
                float(item.subtotal),
                float(item.total),
                txn,
                item.payment_status if item.transaction_type != "return" else labels["none"],
            ]
        )
    return rows


async def export_register_pdf(
    db: AsyncSession,
    *,
    branch_id: int,
    period_start: date,
    period_end: date,
    locale: SalesLocale = "ar",
    limit: int | None = None,
    offset: int = 0,
) -> tuple[bytes, str]:
    labels = _labels(locale)
    align = _cell_align(locale)
    start_inclusive = datetime.combine(period_start, datetime.min.time()).replace(tzinfo=UTC)
    end_exclusive = datetime.combine(period_end + timedelta(days=1), datetime.min.time()).replace(
        tzinfo=UTC,
    )
    items, branch_name = await _register_rows_all(
        db,
        branch_id=branch_id,
        start_inclusive=start_inclusive,
        end_exclusive=end_exclusive,
        limit=limit,
        offset=offset,
    )

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    family = _register_unicode_font(pdf)
    _draw_pdf_logo_banner(pdf, locale=locale, family=family, labels=labels)
    pdf.set_font(family, size=12)
    pdf.cell(0, 8, _txt(labels["register_title"], 120), align=align, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(family, size=8)
    pdf.cell(
        0,
        5,
        _txt(f"{labels['branch']}: {branch_name or branch_id}", 80),
        align=align,
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.cell(
        0,
        5,
        _txt(f"{labels['period']}: {period_start.isoformat()} - {period_end.isoformat()}", 80),
        align=align,
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.ln(2)

    headers = [
        labels["invoice_number"],
        labels["customer"],
        labels["date"],
        labels["time"],
        labels["subtotal"],
        labels["total"],
        labels["txn_type"],
        labels["payment_status"],
    ]
    col_w = [32, 38, 22, 16, 22, 22, 20, 24]
    if locale == "ar":
        headers = list(reversed(headers))
        col_w = list(reversed(col_w))

    pdf.set_font(family, size=7)
    for h, w in zip(headers, col_w, strict=True):
        pdf.cell(w, 6, _txt(h, 24), border=1, align=align)
    pdf.ln()

    table_rows = _register_table_rows(items, locale)
    for row in table_rows:
        if locale == "ar":
            row = list(reversed(row))
        for val, w in zip(row, col_w, strict=True):
            pdf.cell(w, 6, _txt(val, 48), border=1, align=align)
        pdf.ln()

    filename = f"sales-register-{branch_id}-{period_start.isoformat()}-{period_end.isoformat()}.pdf"
    return bytes(pdf.output()), filename


async def export_register_xlsx(
    db: AsyncSession,
    *,
    branch_id: int,
    period_start: date,
    period_end: date,
    locale: SalesLocale = "ar",
    limit: int | None = None,
    offset: int = 0,
) -> tuple[bytes, str]:
    labels = _labels(locale)
    start_inclusive = datetime.combine(period_start, datetime.min.time()).replace(tzinfo=UTC)
    end_exclusive = datetime.combine(period_end + timedelta(days=1), datetime.min.time()).replace(
        tzinfo=UTC,
    )
    items, branch_name = await _register_rows_all(
        db,
        branch_id=branch_id,
        start_inclusive=start_inclusive,
        end_exclusive=end_exclusive,
        limit=limit,
        offset=offset,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = labels["register_title"][:31]
    configure_sheet_locale(ws, locale)
    append_meta_rows(
        ws,
        [
            (labels["register_title"], ""),
            (labels["branch"], branch_name or str(branch_id)),
            (labels["period"], f"{period_start.isoformat()} - {period_end.isoformat()}"),
        ],
        locale=locale,
        title_row=True,
    )
    ws.append([])
    write_table(
        ws,
        [
            labels["invoice_number"],
            labels["customer"],
            labels["date"],
            labels["time"],
            labels["subtotal"],
            labels["total"],
            labels["txn_type"],
            labels["payment_status"],
        ],
        _register_table_rows(items, locale),
        locale=locale,
    )
    filename = f"sales-register-{branch_id}-{period_start.isoformat()}-{period_end.isoformat()}.xlsx"
    return workbook_to_bytes(wb), filename


async def export_daily_sales_summary_pdf(
    db: AsyncSession,
    *,
    branch_id: int | None,
    period_start: date,
    period_end: date,
    locale: SalesLocale = "ar",
) -> tuple[bytes, str]:
    labels = _labels(locale)
    align = _cell_align(locale)
    rows = await _daily_summary_rows(db, branch_id=branch_id, period_start=period_start, period_end=period_end)

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    family = _register_unicode_font(pdf)
    pdf.set_font(family, size=12)
    pdf.cell(0, 8, _txt(labels["daily_title"], 120), align=align, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(family, size=8)
    pdf.cell(
        0,
        5,
        _txt(f"{labels['period']}: {period_start.isoformat()} - {period_end.isoformat()}", 80),
        align=align,
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.ln(2)

    headers = [labels["day"], labels["count"], labels["sales_total"], labels["returns_total"], labels["net_total"]]
    col_w = [36, 24, 36, 36, 36]
    if locale == "ar":
        headers = list(reversed(headers))
        col_w = list(reversed(col_w))

    pdf.set_font(family, size=8)
    for h, w in zip(headers, col_w, strict=True):
        pdf.cell(w, 6, _txt(h, 24), border=1, align=align)
    pdf.ln()
    for row in rows:
        if locale == "ar":
            row = list(reversed(row))
        for val, w in zip(row, col_w, strict=True):
            pdf.cell(w, 6, _txt(val, 48), border=1, align=align)
        pdf.ln()

    filename = f"daily-sales-{period_start.isoformat()}-{period_end.isoformat()}.pdf"
    return bytes(pdf.output()), filename


async def export_daily_sales_summary_xlsx(
    db: AsyncSession,
    *,
    branch_id: int | None,
    period_start: date,
    period_end: date,
    locale: SalesLocale = "ar",
) -> tuple[bytes, str]:
    labels = _labels(locale)
    rows = await _daily_summary_rows(db, branch_id=branch_id, period_start=period_start, period_end=period_end)
    wb = Workbook()
    ws = wb.active
    ws.title = labels["daily_title"][:31]
    configure_sheet_locale(ws, locale)
    append_meta_rows(
        ws,
        [
            (labels["daily_title"], ""),
            (labels["period"], f"{period_start.isoformat()} - {period_end.isoformat()}"),
        ],
        locale=locale,
        title_row=True,
    )
    ws.append([])
    write_table(
        ws,
        [labels["day"], labels["count"], labels["sales_total"], labels["returns_total"], labels["net_total"]],
        rows,
        locale=locale,
    )
    filename = f"daily-sales-{period_start.isoformat()}-{period_end.isoformat()}.xlsx"
    return workbook_to_bytes(wb), filename


async def _daily_summary_rows(
    db: AsyncSession,
    *,
    branch_id: int | None,
    period_start: date,
    period_end: date,
) -> list[list[object]]:
    start_inclusive = datetime.combine(period_start, datetime.min.time()).replace(tzinfo=UTC)
    end_exclusive = datetime.combine(period_end + timedelta(days=1), datetime.min.time()).replace(
        tzinfo=UTC,
    )
    inv_stmt = (
        select(
            func.date(SalesInvoice.created_at).label("day"),
            func.count(SalesInvoice.id).label("cnt"),
            func.coalesce(func.sum(SalesInvoice.total), 0).label("sales_total"),
        )
        .where(
            SalesInvoice.created_at >= start_inclusive,
            SalesInvoice.created_at < end_exclusive,
            SalesInvoice.voided_at.is_(None),
        )
        .group_by(func.date(SalesInvoice.created_at))
    )
    if branch_id is not None:
        inv_stmt = inv_stmt.where(SalesInvoice.branch_id == branch_id)

    ret_stmt = (
        select(
            func.date(CreditNote.created_at).label("day"),
            func.coalesce(func.sum(CreditNote.total_amount), 0).label("returns_total"),
        )
        .join(SalesReturn, CreditNote.sales_return_id == SalesReturn.id)
        .join(SalesInvoice, SalesReturn.sales_invoice_id == SalesInvoice.id)
        .where(
            CreditNote.created_at >= start_inclusive,
            CreditNote.created_at < end_exclusive,
        )
        .group_by(func.date(CreditNote.created_at))
    )
    if branch_id is not None:
        ret_stmt = ret_stmt.where(SalesInvoice.branch_id == branch_id)

    sales_by_day: dict[str, tuple[int, Decimal]] = {}
    for row in (await db.execute(inv_stmt)).all():
        sales_by_day[str(row.day)] = (int(row.cnt or 0), Decimal(str(row.sales_total or 0)))

    returns_by_day: dict[str, Decimal] = {}
    for row in (await db.execute(ret_stmt)).all():
        returns_by_day[str(row.day)] = Decimal(str(row.returns_total or 0))

    all_days = sorted(set(sales_by_day.keys()) | set(returns_by_day.keys()))
    out: list[list[object]] = []
    for day in all_days:
        cnt, sales_total = sales_by_day.get(day, (0, Decimal("0")))
        returns_total = returns_by_day.get(day, Decimal("0"))
        net = sales_total - returns_total
        out.append([day, cnt, float(sales_total), float(returns_total), float(net)])
    return out
